"""Benchmark runner and Excel export."""

from __future__ import annotations

import statistics
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import tiktoken
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from multi_llm_speed_comparison.clients import build_client
from multi_llm_speed_comparison.config import (
    BENCHMARK_TASKS,
    MODEL_CONFIGS,
    RUNS_PER_MODEL,
    TEMPERATURE,
    BenchmarkTask,
    ModelConfig,
)


@dataclass(frozen=True)
class RunDetail:
    run_number: int
    response_time_seconds: float | None
    output_tokens: int | None
    tokens_per_second: float | None
    answer: str
    error: str | None = None


@dataclass(frozen=True)
class TaskAverage:
    response_time_seconds: float | None
    output_tokens: float | None
    tokens_per_second: float | None
    details: list[RunDetail]
    error: str | None = None


def run_benchmark(output_path: Path) -> Path:
    load_dotenv()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    results: dict[str, dict[str, TaskAverage]] = {}
    for model_config in MODEL_CONFIGS:
        results[model_config.display_name] = {}
        for task in BENCHMARK_TASKS:
            results[model_config.display_name][task.name] = _run_task_average(
                model_config=model_config,
                task=task,
                runs=RUNS_PER_MODEL,
            )

    _write_excel(output_path=output_path, results=results)
    return output_path


def default_output_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path("outputs") / f"llm_evaluation_{timestamp}.xlsx"


def _run_task_average(
    model_config: ModelConfig,
    task: BenchmarkTask,
    runs: int,
) -> TaskAverage:
    try:
        client = build_client(model_config)
    except Exception as exc:  # noqa: BLE001
        return TaskAverage(None, None, None, [], str(exc))

    latencies: list[float] = []
    output_token_counts: list[int] = []
    throughputs: list[float] = []

    details: list[RunDetail] = []

    for run_number in range(1, runs + 1):
        try:
            started_at = time.perf_counter()
            response = client.complete(task.prompt)
            latency = time.perf_counter() - started_at
            
            # Some providers don't return output tokens, so we need to estimate them
            if response.output_tokens is not None:
                output_tokens_real = response.output_tokens
                output_tokens_estimate = _estimate_tokens(response.text)
                output_tokens = output_tokens_real
                # print(f"output_tokens_real: {output_tokens_real}, output_tokens_estimate: {output_tokens_estimate}, model: {model_config.display_name}")   
            else:
                output_tokens = _estimate_tokens(response.text)
                print(f"output_tokens_real: None, output_tokens_estimate: {output_tokens}, model: {model_config.display_name}")
            token_per_second = output_tokens / latency if latency > 0 else 0
        except Exception as exc:  # noqa: BLE001
            error = str(exc)
            details.append(
                RunDetail(
                    run_number=run_number,
                    response_time_seconds=None,
                    output_tokens=None,
                    tokens_per_second=None,
                    answer="",
                    error=error,
                )
            )
            return TaskAverage(None, None, None, details, error)

        latencies.append(latency)
        output_token_counts.append(output_tokens)
        throughputs.append(token_per_second)
        details.append(
            RunDetail(
                run_number=run_number,
                response_time_seconds=latency,
                output_tokens=output_tokens,
                tokens_per_second=token_per_second,
                answer=response.text,
            )
        )

    return TaskAverage(
        response_time_seconds=statistics.fmean(latencies),
        output_tokens=statistics.fmean(output_token_counts),
        tokens_per_second=statistics.fmean(throughputs),
        details=details,
    )


def _estimate_tokens(text: str) -> int:
    if not text:
        return 0
    try:
        encoding = tiktoken.get_encoding("cl100k_base")
        return len(encoding.encode(text))
    except Exception:  # noqa: BLE001
        return max(1, len(text) // 4)


def _write_excel(
    output_path: Path,
    results: dict[str, dict[str, TaskAverage]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    _write_summary_header(sheet)

    for model_name, task_results in results.items():
        row: list[object] = [model_name]
        errors: list[str] = []
        for task in BENCHMARK_TASKS:
            average = task_results[task.name]
            row.append(_round_or_blank(average.response_time_seconds))
            row.append(_round_or_blank(average.output_tokens))
            row.append(_round_or_blank(average.tokens_per_second))
            if average.error:
                errors.append(f"{task.name}: {average.error}")
        row.append("; ".join(errors))
        sheet.append(row)

    _set_summary_column_widths(sheet)

    _write_details_sheet(workbook=workbook, results=results)
    _write_tasks_sheet(workbook=workbook)

    workbook.save(output_path)


def _write_summary_header(sheet) -> None:
    sheet.cell(row=1, column=1, value="Model - Provider")
    sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    column_index = 2
    for task in BENCHMARK_TASKS:
        sheet.cell(row=1, column=column_index, value=task.name)
        sheet.merge_cells(
            start_row=1,
            start_column=column_index,
            end_row=1,
            end_column=column_index + 2,
        )
        sheet.cell(row=2, column=column_index, value="response time")
        sheet.cell(row=2, column=column_index + 1, value="output tokens")
        sheet.cell(row=2, column=column_index + 2, value="token/second")
        column_index += 3

    sheet.cell(row=1, column=column_index, value="error")
    sheet.merge_cells(
        start_row=1,
        start_column=column_index,
        end_row=2,
        end_column=column_index,
    )

    for row in sheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_summary_column_widths(sheet) -> None:
    max_column = sheet.max_column
    sheet.column_dimensions["A"].width = 42
    for column_index in range(2, max_column):
        column_letter = sheet.cell(row=2, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = 16
    error_column = sheet.cell(row=1, column=max_column).column_letter
    sheet.column_dimensions[error_column].width = 60


def _write_tasks_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Tasks")
    sheet.append(["Name", "Prompt", "Temperature"])

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for task in BENCHMARK_TASKS:
        sheet.append([task.name, task.prompt, TEMPERATURE])

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 100
    sheet.column_dimensions["C"].width = 16


def _write_details_sheet(
    workbook: Workbook,
    results: dict[str, dict[str, TaskAverage]],
) -> None:
    sheet = workbook.create_sheet("Details")
    sheet.append(
        [
            "Model",
            "Task",
            "Run",
            "Response time",
            "Output tokens",
            "Token/second",
            "Answer",
            "Error",
        ]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for model_name, task_results in results.items():
        for task in BENCHMARK_TASKS:
            average = task_results[task.name]
            if not average.details:
                sheet.append(
                    [model_name, task.name, "", "", "", "", "", average.error or ""]
                )
                continue

            for detail in average.details:
                sheet.append(
                    [
                        model_name,
                        task.name,
                        detail.run_number,
                        _round_or_blank(detail.response_time_seconds),
                        detail.output_tokens
                        if detail.output_tokens is not None
                        else "",
                        _round_or_blank(detail.tokens_per_second),
                        detail.answer,
                        detail.error or "",
                    ]
                )

    column_widths = {
        "A": 36,
        "B": 12,
        "C": 8,
        "D": 16,
        "E": 16,
        "F": 16,
        "G": 80,
        "H": 80,
    }
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width


def _round_or_blank(value: float | None) -> float | str:
    if value is None:
        return ""
    return round(value, 4)
